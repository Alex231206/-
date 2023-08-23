import openpyxl
import os
from os.path import isabs
from datetime import datetime
from openpyxl.utils.cell import get_column_letter, column_index_from_string
from time import time

def print_annotation():
    """
    This program takes xlsx file (the name of it is entered by user) and turns columns into rows and rows into columns.
    It creates a new sheet in the file
    New sheet name pattern is 'Transformed sheet_16_48_2"""

def print_command_list():
    """
    1) 0 - to quit currently running program
    2) 1 - to resume currently running program and move on
    3) 2 - to show the annotation
    4) 3 - to return to the beginning"""

while True:
    print(print_command_list.__doc__)

    user_choice = input('Enter the number of command: ')

    if user_choice == '0':
        print('See you soon!')
        break

    elif user_choice == '1':
        abs_path_to_folder = input('Enter the absoulute path to the folder which contains xlsx file: ')

        if isabs(abs_path_to_folder):
            os.chdir(abs_path_to_folder)

            filename = input('Enter the name of xlsx file: ')
            wb = openpyxl.load_workbook(filename)

            if filename in os.listdir(abs_path_to_folder) and filename[filename.rfind('.') + 1:] == 'xlsx':
                basic_sheet_name = input('Enter the name of xlsx file sheet which contains data you want to get transformed: ')

                start = time()

                if basic_sheet_name in wb.sheetnames:
                    date = datetime.today()

                    new_sheet_title = f'Transformed sheet {date.hour}_{date.minute}_{date.second}'

                    wb.create_sheet(title = new_sheet_title)

                    basic_sheet = wb[basic_sheet_name]
                    new_sheet = wb[new_sheet_title]

                    first_cell_coordinate = 'A1'
                    last_cell_coordinate = f'{get_column_letter(basic_sheet.max_column)}{basic_sheet.max_row}'

                    for row in basic_sheet[first_cell_coordinate: last_cell_coordinate]:
                        for cell in row:
                            if cell.row == 1:
                                new_sheet[f'A{cell.column}'] = cell.value

                            else:
                                new_sheet[f'{get_column_letter(cell.row)}{cell.column}'] = cell.value

                    wb.save(filename)

                    print(f'New sheet with transformed data has been successfully created. It took {time() - start} seconds\n')
                    continue

                else:
                    print(f"Sheet {basic_sheet_name} doesn't exist\n")
                    continue

            else:
                print(f"File {filename} doesn't exist or its extension doesn't match xlsx extension\n")
                continue

        else:
            print(f"The path {abs_path_to_folder} doesn't exist\n")
            continue

    elif user_choice == '2':
        print(print_annotation.__doc__)
        continue

    elif user_choice == '3':
        continue

    else:
        print(f'Command number {user_choice} is incorrect\n')
        continue

